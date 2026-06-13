#!/usr/bin/env python3
"""Convert all Excel data to JSON for React app"""
import json
import openpyxl

def convert_noip():
    wb = openpyxl.load_workbook('/Users/zengyiqing/Documents/test/ssoi/NOIP总(1988-2024).xlsx', data_only=True)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None or not isinstance(row[0], int):
            continue
        data.append({
            'year': row[0],
            '提高组': {
                '一等奖': row[1] or 0,
                '二等奖': row[2] or 0,
                '三等奖': row[3] or 0,
                '获奖人次': row[4] or 0
            },
            '普及组': {
                '一等奖': row[5] or 0,
                '二等奖': row[6] or 0,
                '三等奖': row[7] or 0,
                '获奖人次': row[8] or 0
            }
        })
    return data

def convert_csp():
    wb = openpyxl.load_workbook('/Users/zengyiqing/Documents/test/ssoi/CSP2019-2024汇总.xlsx', data_only=True)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None or not isinstance(row[0], int):
            continue
        data.append({
            'year': row[0],
            'CSP-S': {
                '一等奖': row[1] or 0,
                '二等奖': row[2] or 0,
                '三等奖': row[3] or 0,
                '获奖人次': row[4] or 0
            },
            'CSP-J': {
                '一等奖': row[5] or 0,
                '二等奖': row[6] or 0,
                '三等奖': row[7] or 0,
                '获奖人次': row[8] or 0
            }
        })
    return data

def convert_apio():
    wb = openpyxl.load_workbook('/Users/zengyiqing/Documents/test/ssoi/2010-2024APIO奖牌统计.xlsx', data_only=True)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None or not isinstance(row[0], int):
            continue
        data.append({
            'year': row[0],
            '金牌': row[1] or 0,
            '银牌': row[2] or 0,
            '铜牌': row[3] or 0,
            '总数': row[4] or 0
        })
    return data

def convert_noi():
    wb = openpyxl.load_workbook('/Users/zengyiqing/Documents/test/ssoi/信息学全国赛成绩统计.xlsx', data_only=True)

    data = []
    # Only process Sheet1 which has all the data
    ws = wb['Sheet1']
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None or not isinstance(row[0], int):
            continue
        entry = {
            'id': row[0],
            'name': row[1],
            'competition': row[2],
            'level': row[3],
            'rank': row[4]
        }
        if len(row) > 5 and row[5]:
            entry['school'] = row[5]
        data.append(entry)
    return data

# Convert all data
noip_data = convert_noip()
csp_data = convert_csp()
apio_data = convert_apio()
noi_data = convert_noi()

# Save to JSON files
with open('/Users/zengyiqing/Documents/test/ssoi-web/src/data/noip.json', 'w', encoding='utf-8') as f:
    json.dump(noip_data, f, ensure_ascii=False, indent=2)

with open('/Users/zengyiqing/Documents/test/ssoi-web/src/data/csp.json', 'w', encoding='utf-8') as f:
    json.dump(csp_data, f, ensure_ascii=False, indent=2)

with open('/Users/zengyiqing/Documents/test/ssoi-web/src/data/apio.json', 'w', encoding='utf-8') as f:
    json.dump(apio_data, f, ensure_ascii=False, indent=2)

with open('/Users/zengyiqing/Documents/test/ssoi-web/src/data/noi.json', 'w', encoding='utf-8') as f:
    json.dump(noi_data, f, ensure_ascii=False, indent=2)

print(f"NOIP data: {len(noip_data)} years")
print(f"CSP data: {len(csp_data)} years")
print(f"APIO data: {len(apio_data)} years")
print(f"NOI data: {len(noi_data)} entries")
print("All data converted successfully!")